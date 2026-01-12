# core/admin.py
from __future__ import annotations

from typing import Any

from django.contrib import admin
from django.contrib.admin.utils import label_for_field, lookup_field
from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.http import HttpRequest, HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import formats

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .forms import AdminExportForm
from .models import Category, Event, VolunteerApplication, EventLike


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "created_at", "updated_at")
    search_fields = ("name",)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ("id", "title", "category", "event_date", "location", "created_at")
    list_filter = ("category",)
    search_fields = ("title", "location")


@admin.register(VolunteerApplication)
class VolunteerApplicationAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "event", "status", "created_at")
    list_filter = ("status", "event")
    search_fields = ("user__username", "event__title")


@admin.register(EventLike)
class EventLikeAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "event", "created_at")
    search_fields = ("user__username", "event__title")


def _format_admin_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "Да" if value else "Нет"

    if isinstance(value, models.Model):
        return str(value)

    # datetime/date: локализованный формат как в админке
    if hasattr(value, "strftime"):
        try:
            return formats.localize(value)
        except Exception:
            return str(value)

    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(x) for x in value)

    return value


def _get_admin_columns_and_headers(
    request: HttpRequest,
    model_admin: admin.ModelAdmin,
) -> tuple[list[str], list[str]]:
    """
    Базовые колонки (как в list_display) + заголовки.
    """
    columns = list(model_admin.get_list_display(request))
    headers: list[str] = []

    for col in columns:
        try:
            header = label_for_field(col, model_admin.model, model_admin=model_admin, return_attr=False)
        except Exception:
            header = col
        headers.append(str(header))

    return columns, headers


def _get_fields_choices_for_model(
    request: HttpRequest,
    model_admin: admin.ModelAdmin,
) -> tuple[list[tuple[str, str]], list[str]]:
    """
    Возвращает:
    - choices для чекбоксов: (column_name, header)
    - список всех column_name (для initial)
    """
    columns, headers = _get_admin_columns_and_headers(request, model_admin)
    choices = list(zip(columns, headers))
    return choices, columns


def _get_admin_row_values(
    model_admin: admin.ModelAdmin,
    obj: models.Model,
    columns: list[str],
) -> list[Any]:
    row: list[Any] = []

    for col in columns:
        try:
            _, _, value = lookup_field(col, obj, model_admin)
        except ObjectDoesNotExist:
            value = ""
        except Exception:
            value = ""

        row.append(_format_admin_value(value))

    return row


def export_xlsx_view(request: HttpRequest) -> HttpResponse:
    """
    Экспорт XLSX:
    - выбираем одну модель
    - выбираем набор колонок (list_display)
    """
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)

    # текущие ModelAdmin из стандартного admin.site
    model_admin_map: dict[str, admin.ModelAdmin] = {
        "core.Category": admin.site._registry.get(Category),
        "core.Event": admin.site._registry.get(Event),
        "core.VolunteerApplication": admin.site._registry.get(VolunteerApplication),
        "core.EventLike": admin.site._registry.get(EventLike),
    }
    model_admin_map = {k: v for k, v in model_admin_map.items() if v is not None}

    # --- Формирование формы ---
    form = AdminExportForm(request.POST or None)
    form.fields["model"].choices = [(k, k) for k in model_admin_map.keys()]

    selected_model_label = None
    selected_model_admin: admin.ModelAdmin | None = None

    # берём выбранную модель (из POST, если есть)
    if request.method == "POST":
        selected_model_label = request.POST.get("model") or None
    else:
        # на случай, если захочешь потом сделать выбор через GET — можно расширить
        selected_model_label = None

    if selected_model_label and selected_model_label in model_admin_map:
        selected_model_admin = model_admin_map[selected_model_label]

        # заполняем choices для полей
        field_choices, all_columns = _get_fields_choices_for_model(request, selected_model_admin)
        form.fields["fields"].choices = field_choices

        # если это первый показ после выбора модели — по умолчанию отмечаем все
        if request.method != "POST" or ("preview" in request.POST and not request.POST.getlist("fields")):
            form.initial["fields"] = all_columns

    # --- POST обработка ---
    if request.method == "POST" and form.is_valid():
        selected_model_label = form.cleaned_data["model"]
        selected_model_admin = model_admin_map.get(selected_model_label)

        if selected_model_admin is None:
            return TemplateResponse(request, "admin/export_xlsx.html", {"form": form})

        # права на просмотр модели
        app_label = selected_model_admin.model._meta.app_label
        model_name = selected_model_admin.model._meta.model_name
        perm = f"{app_label}.view_{model_name}"
        if not (request.user.is_superuser or request.user.has_perm(perm)):
            return HttpResponse("Forbidden", status=403)

        # если нажали "Показать поля" — просто перерендерим форму (без скачивания)
        if "preview" in request.POST and "download" not in request.POST:
            # подсветить все поля по умолчанию (если ещё не выбраны)
            if not form.cleaned_data.get("fields"):
                _, all_columns = _get_fields_choices_for_model(request, selected_model_admin)
                form.initial["fields"] = all_columns
            return TemplateResponse(request, "admin/export_xlsx.html", {"form": form})

        # если нажали "Скачать"
        selected_fields: list[str] = form.cleaned_data.get("fields") or []
        all_columns, all_headers = _get_admin_columns_and_headers(request, selected_model_admin)

        if selected_fields:
            # оставляем только выбранные, сохраняя порядок list_display
            columns = [c for c in all_columns if c in set(selected_fields)]
        else:
            # если ничего не выбрано — выгружаем все
            columns = all_columns

        # заголовки под выбранные колонки
        headers_map = dict(zip(all_columns, all_headers))
        headers = [headers_map.get(c, c) for c in columns]

        wb = Workbook()
        ws = wb.active
        ws.title = selected_model_label.split(".")[-1][:31]

        ws.append(headers)

        qs = selected_model_admin.get_queryset(request).order_by("id")[:5000]
        for obj in qs:
            ws.append(_get_admin_row_values(selected_model_admin, obj, columns))

        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(45, len(str(header)) + 6))

        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="export.xlsx"'
        wb.save(resp)
        return resp

    return TemplateResponse(request, "admin/export_xlsx.html", {"form": form})


# Просто добавляем URL в существующий admin.site через обёртку.
_original_get_urls = admin.site.get_urls


def _patched_get_urls():
    urls = _original_get_urls()
    custom = [
        path("export-xlsx/", admin.site.admin_view(export_xlsx_view), name="export_xlsx"),
    ]
    return custom + urls


admin.site.get_urls = _patched_get_urls  # type: ignore[method-assign]
